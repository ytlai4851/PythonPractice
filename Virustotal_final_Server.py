# -*- coding : utf-8 -*-
import simplejson, urllib, urllib2, ctypes, collections, MySQLdb, time, os
from openpyxl import load_workbook, Workbook
from openpyxl.chart import BarChart, PieChart, Reference


class DB:
    db = MySQLdb.connect("IP", "DBUser", "passed", "virusdb", charset="utf8")

    # db=MySQLdb.connect("127.0.0.1","root","!QAZ2wsx","virusdb",charset="utf8")
    cursor = db.cursor()

    def MD5InforSearch(slef, MD5Value):  # Creat
        infor = []
        SQLCommand = "SELECT * FROM data WHERE MD5 = %s "
        slef.cursor.execute(SQLCommand, (MD5Value,))
        rows = slef.cursor.fetchall()
        for i in rows:
            infor = [i[0], i[1], i[2], i[3], i[4]]
        return (infor)

    def UpdateDBMD5(slef, MD5, VTP, VTT, Reslut, VTURL, From, CurrentTime):
        print(MD5, VTP, VTT, Reslut, VTURL, From, CurrentTime)
        SQLCommand = "INSERT INTO data(MD5,VTPositives,VTTotal,Result,VTURL,UpdateTime) VALUES (%s,%s,%s,%s,%s,%s)"
        slef.cursor.execute(SQLCommand, (MD5, int(VTP), int(VTT), Reslut, VTURL, CurrentTime))
        slef.db.commit()
        slef.MD5InforSearch(MD5)


def CreatMalwareWorkbook(WorkbookNmae, SheetName, Title):
    AddSheet = WorkbookNmae.create_sheet(title=SheetName)
    AddSheet.append(Title)  # Excel header
    return AddSheet


class VTCheck():

    def __init__(slef, HashValue):
        slef.HashValue = HashValue

    def VirustotalSerch(slef):
        url = "https://www.virustotal.com/vtapi/v2/file/report"
        parameters = {"resource": slef.HashValue,
                      "apikey": "zzzzzzzzzzz"}
        data = urllib.urlencode(parameters)
        req = urllib2.Request(url, data)
        try:
            response = urllib2.urlopen(req)
        except:
            print("Plz Check ur Network")
            return
        json = response.read()
        Answer = simplejson.loads(json)
        time.sleep(16)
        return Answer

    def AnalysisMD5(slef):
        TestMode = 0
        print("Search: " + slef.HashValue)
        if TestMode == 0:
            Answer = slef.VirustotalSerch()

        else:
            Answer = {"6d87ee730e4a2a20782e9cbb10a106ac": [1, 57,
                                                           "https://www.virustotal.com/file/076b51bdefd5b50f30b25b00c26233127f5819bac97de72177cab1e82f90da16/analysis/1473062036/",
                                                           "F"], "8f311a272aae611bffaac88cc0ca3f43": [0, 57,
                                                                                                      "https://www.virustotal.com/file/11e72607b7caf41360e9797ac871a9d9bfca06d51ce2d75d1cc9ee78c57f25f3/analysis/1474017476/",
                                                                                                      "F"],
                      "aebceffb36ba9f9ac308ef079818f9fd": [0, 57,
                                                           "https://www.virustotal.com/file/630836b98d72f814d0523d82a77eb92464c2fe1c42a97dc59ba1e3a9651df4c0/analysis/1472651966/",
                                                           "F"], "7cf1b716372b89568ae4c0fe769f5869": [0, 57,
                                                                                                      "https://www.virustotal.com/file/0d70a7a594bcfbb26d7249c0f4b0af9ef874f2318b3fdce44648cc61279594ed/analysis/1474005281/",
                                                                                                      "F"],
                      "f0738e835e93d6735adc9f90efd0d78f": [0, 56,
                                                           "https://www.virustotal.com/file/0957c2882c85592c6f1eb766ad0b61be04d19684e7261058e3ab4fd9db93adba/analysis/1473351154/",
                                                           "F"], "1f16ba2f663531d78c166d116c15e087": [0, 54,
                                                                                                      "https://www.virustotal.com/file/ecfa29e6792d9000a482fba46db06c8f6ffd744f3defec82649585e69ed048d0/analysis/1467549629/",
                                                                                                      "F"],
                      "207fd81633c51557ea3efa9edd300b8f": [0, 57,
                                                           "https://www.virustotal.com/file/25aa858371fc181f6cb125ddc89d27aff034ca2bcf324fb898bf3583ca127cd6/analysis/1472547557/",
                                                           "F"], "b63b8e243ff5a1e63940e7f27441b10d": [0, 56,
                                                                                                      "https://www.virustotal.com/file/5bf03a46b84b522b48ec00f170c57cf149462b2de42f9c73936827227aa3ec6e/analysis/1462158531/",
                                                                                                      "F"],
                      "9f264f5455bd09a8736ef2dd0dd9536a": [1, 54,
                                                           "https://www.virustotal.com/file/e7a72436ef442eef9f33a4e4db1dd41623964ee84261d8e22398c402f2b47656/analysis/1413982782/",
                                                           "F"], "9f13849d6734e14278e6b03da1fd9b90": [0, 54,
                                                                                                      "https://www.virustotal.com/file/05a8ea16c1ff4f538c9a0b0ab9ab036f11448ad53b07cffc4203a88c6caa8007/analysis/1411453969/",
                                                                                                      "F"],
                      "8f9dabffebfbe6bab664626fd84d4d39": [0, 56,
                                                           "https://www.virustotal.com/file/40d7060f4b3a9adecf86c58459aeff39d73dac2e8a5a56467b96f2b5e324e805/analysis/1465507165/",
                                                           "F"],
                      "a334c5240bde5a673e73b0a6a5be293c": ["NA", "NA", "NA", "C"],
                      "1bbf41fa584cd0ae880e86b861c9d7be": ["NA", "NA", "NA", "C"],
                      "959217ca3f2ad82281ca3b811801b110": [0, 56,
                                                           "https://www.virustotal.com/file/fc5664b0afaf7612ca69187519ecdfec5c748dc1cbf32d3e2a03ca85429a6168/analysis/1473831405/",
                                                           "F"], "07d61a6f4688e5cb37b0d2e379440f24": [1, 56,
                                                                                                      "https://www.virustotal.com/file/ded2d5c79612c69c892c60e929dd250a85a0bcb0ad5875c7c151e5f24f031960/analysis/1464861793/",
                                                                                                      "F"],
                      "f67e9550ee84a413a84c43725085ed1d": ["NA", "NA", "NA", "C"],
                      "b2428543c7597224916ecd47f8ef07ea": ["NA", "NA", "NA", "C"],
                      "d567a6d0647f80ecb5a761ddd9ad367c": [1, 58,
                                                           "https://www.virustotal.com/file/fa76d9059fc4fe103e3de1868c76d86efbac9295fcd3e5f1c430cd1c5b4ad710/analysis/1472543886/",
                                                           "F"], "2dff216851d0b4bbd1e452942d7499c0": [18, 56,
                                                                                                      "https://www.virustotal.com/file/766e5db40f1b5672131ae0f5f4052396b96743381c74ed86705c7692d8c4fab9/analysis/1463714919/",
                                                                                                      "T"],
                      "f0773eaf972083bb03e92fea13bb961d": [0, 58,
                                                           "https://www.virustotal.com/file/44c98c103e24a6f3fa3543a90c91ac212389c30c4919a07a95489573215338b5/analysis/1472621116/",
                                                           "F"],
                      "05ad16d04d79886fbb6ea9ceb44cb71c": ["NA", "NA", "NA", "C"]}
        try:
            positivesRate = float(Answer["positives"]) / float(Answer["total"])
            if positivesRate > 0.1:
                positivesRate = "T"
            else:
                positivesRate = "F"
            d.UpdateDBMD5(Answer["md5"], Answer["positives"], Answer["total"], positivesRate, str(Answer["permalink"]),
                          "VST", time.strftime("%Y-%m-%d"))
        except:
            if len(slef.HashValue) == 32:
                d.UpdateDBMD5(slef.HashValue, 0, 0, "C", "NA", "NA",
                              time.strftime("%Y-%m-%d"))  # virustotal can't find md5


class AnalysisData():

    def __init__(self, CreateFinalWorkbook):
        self.CreateFinalWorkbook = CreateFinalWorkbook

    def FileStatistics(self, LevelFileCount, FileName):  # make chart
        FileStatisticsSheet = WorkSheet = CreatMalwareWorkbook(CreateFinalWorkbook, FileName, ["Filename", "Count"])
        for key, value in dict(collections.Counter(LevelFileCount)).iteritems():  # get filename and filenamecount
            FileStatisticsSheet.append([key, value])
        MD5Chart = BarChart()
        data = Reference(FileStatisticsSheet, min_col=2, min_row=2,
                         max_row=len(collections.Counter(LevelFileCount)) + 1)  # excel value reference
        cat = Reference(FileStatisticsSheet, min_col=1, min_row=2,
                        max_row=len(collections.Counter(LevelFileCount)) + 1)  # bar categories reference
        MD5Chart.add_data(data)
        MD5Chart.set_categories(cat)
        FileStatisticsSheet.add_chart(MD5Chart, "C1")

    def MalWareAnalysis(self, FileName):
        L3FilenameCount = []
        L45FilenameCount = []
        suspiciousComputer = set()
        FinshMd5 = []
        NeedReChackSheet = CreatMalwareWorkbook(CreateFinalWorkbook, "%s_NeedReChack" % (FileName),
                                                ["ComputerName", "MFileName", "MD5Value"])
        WorkSheet = CreatMalwareWorkbook(CreateFinalWorkbook, FileName,
                                         ["Filename", "Company Name", "Reslut", "Computer Name"])
        for j in xrange(1, len(LoadMalwareFindingsSheet.rows)):  # read rows in malwarefind excel
            Level = LoadMalwareFindingsSheet.cell(row=j + 1, column=1).value
            MFileName = LoadMalwareFindingsSheet.cell(row=j + 1, column=2).value
            ComputerName = LoadMalwareFindingsSheet.cell(row=j + 1, column=3).value
            MCompany = LoadMalwareFindingsSheet.cell(row=j + 1, column=4).value
            MD5Value = LoadMalwareFindingsSheet.cell(row=j + 1, column=5).value

            if MD5Value not in FinshMd5:
                FinshMd5.append(MD5Value)
                md5_dict = d.MD5InforSearch(MD5Value)

                if md5_dict == []:
                    VT = VTCheck(MD5Value)
                    VT.AnalysisMD5()
                    md5_dict = d.MD5InforSearch(MD5Value)
                print(md5_dict)
                WorkSheet.append(
                    [MFileName.split("\\")[-1], MCompany, md5_dict[3], ComputerName, md5_dict[1], md5_dict[2],
                     md5_dict[4]])  # write to excel
            if md5_dict[3] != "F":
                NeedReChackSheet.append([ComputerName, MFileName, MD5Value])
            if md5_dict[3] == "T":
                suspiciousComputer.update([ComputerName])
            if Level == 3:
                L3FilenameCount.append(MFileName.split("\\")[-1])
            else:
                L45FilenameCount.append(MFileName.split("\\")[-1])

        if len(L3FilenameCount) > 0:
            self.FileStatistics(L3FilenameCount, "%s_L3FileStat" % (FileName))
        if len(L45FilenameCount) > 0:
            self.FileStatistics(L45FilenameCount, "%s_L45FileStat" % (FileName))

        CreateFinalWorkbook.save("ReportInfo.xlsx")
        return suspiciousComputer

    def LevelStatistic(self):
        # print(Workbook)
        LevelCount = [0, 0, 0, 0, 0]
        LevelStatisticSheet = CreatMalwareWorkbook(CreateFinalWorkbook, "LevelStatistic", ["Level", "Count"])
        for i in xrange(1, len(AllDevicesSheet.rows)):
            ComputerLevel = int(AllDevicesSheet.cell(row=i + 1, column=6).value) - 1
            if ComputerLevel >= 0:
                LevelCount[ComputerLevel] += 1

        for i in xrange(len(LevelCount)):  # for make pei chart
            LevelStatisticSheet.append(["Level %d" % (i + 1), LevelCount[i]])

        for i in xrange(len(LevelCount)):  # for word table
            LevelStatisticSheet.cell(row=10, column=i + 1).value = "Level %d" % (i + 1)
            LevelStatisticSheet.cell(row=11, column=i + 1).value = str(LevelCount[i]) + "( %s )" % "{:.0%}".format(
                LevelCount[i] / float(sum(LevelCount)))

        LevelPeiChart = PieChart()
        cat = Reference(LevelStatisticSheet, min_row=1, min_col=1, max_row=5)
        data = Reference(LevelStatisticSheet, min_row=1, min_col=2, max_row=5)
        LevelPeiChart.add_data(data)
        LevelPeiChart.set_categories(cat)
        LevelStatisticSheet.add_chart(LevelPeiChart, "G1")
        CreateFinalWorkbook.save("ReportInfo.xlsx")

    def OsStatistic(self, suspici):
        OsP = []  # suspicious
        OsT = []
        OsStatisticSheet = CreatMalwareWorkbook(CreateFinalWorkbook, "OsStatistic",
                                                ["Os Version", "NonSuspicious", "Suspicious", "Total"])
        for i in xrange(1, len(AllDevicesSheet.rows)):
            if int(AllDevicesSheet.cell(row=i + 1, column=6).value) > 0:  # scan succeed or not
                if AllDevicesSheet.cell(row=i + 1, column=1).value in suspici:
                    OsP.append(AllDevicesSheet.cell(row=i + 1, column=2).value)
                OsT.append(AllDevicesSheet.cell(row=i + 1, column=2).value)  # os version
        for key, value in dict(collections.Counter(OsT)).iteritems():  # get filename and filenamecount
            OsStatisticSheet.append([key, value - collections.Counter(OsP)[key], collections.Counter(OsP)[key], value])


d = DB()
LoadAllDevices = load_workbook(filename='%s\\MalFile\\XecProbe_AllDevices.xlsx' % (os.getcwd()))
AllDevicesSheet = LoadAllDevices["All Devices"]
CreateFinalWorkbook = Workbook()
AnalysisSheet = AnalysisData(CreateFinalWorkbook)
SplitRawDataCount = []
for i in os.listdir("%s\\MalFile" % (os.getcwd())):
    if i != "XecProbe_AllDevices.xlsx":
        SplitRawDataCount.append(i)
AnalysisSheet.LevelStatistic()

while len(SplitRawDataCount) > 0:
    print(SplitRawDataCount[0])
    LoadMalwareFindings = load_workbook(filename='%s\\MalFile\\%s' % (os.getcwd(), SplitRawDataCount[0]))
    LoadMalwareFindingsSheet = LoadMalwareFindings[LoadMalwareFindings.sheetnames[0]]
    SuspicousComputerName = AnalysisSheet.MalWareAnalysis(SplitRawDataCount[0])
    del SplitRawDataCount[0]

AnalysisSheet.OsStatistic(SuspicousComputerName)
CreateFinalWorkbook.save("ReportInfo.xlsx")
ctypes.windll.user32.MessageBoxA(0, "Done", "", 0)
